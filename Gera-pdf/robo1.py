from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
# biblioteca openpyxl -> para lê do excel
import openpyxl
from fpdf import FPDF

arq = open('result.txt', 'w')
dominios = []

print('Iniciando nosso robô...\n')
driver = webdriver.Chrome('/home/kellina/Projetos/Robos/chromedriver')
driver.get('https://registro.br')

path = '/home/kellina/Projetos/Robos/dominios.xlsx'
book = openpyxl.load_workbook(path)
ws = book.active
m_linha = ws.max_row

for i in range(1, m_linha + 1): # Lẽ da linha 0 a 10
  valores = ws.cell(row = i, column = 1)
  dominios.append(valores.value)
  
for dominio in dominios:
  pesquisa = driver.find_element_by_id('is-avail-field')
  pesquisa.clear()
  pesquisa.send_keys(dominio)
  pesquisa.send_keys(Keys.RETURN)
  time.sleep(2)
  resultados = driver.find_elements_by_tag_name('strong')
  texto = 'Domínio %s %s\n' %(dominio, resultados[4].text)
  arq.write(texto) # Para exrever os resultados no result.txt

arq.close()
driver.close()
print('Finalizado com sucesso!\n')

# Covertendo  de txt para PDF
pdf = FPDF()
pdf.add_page()
pdf.set_font('Arial', size=15)
arq = open('result.txt', 'r')
for x in arq:
  pdf.cell(200, 10, txt = x, ln = 1, align = 'C')
pdf.output('result.pdf')